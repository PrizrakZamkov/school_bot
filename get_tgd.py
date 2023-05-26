from openpyxl import load_workbook
from test_dirs_os import get_all_xlsx


def get_data_students(file_path='student_timetables'):
    tables_name = get_all_xlsx(file_path)
    dict_s = {}
    for i in tables_name:
        book = load_workbook(i)
        sheet = book[book.active.title]
        index = 1
        next_sheet_cell = sheet.cell(row=1, column=index).value
        while next_sheet_cell != None:
            name = str(next_sheet_cell).lower()
            dict_s[name] = []
            for day in range(5):
                data = []
                for f in range(day*9 + 2, day*9 + 11):
                    if sheet.cell(row=f, column=index).value != None:
                        data.append(sheet.cell(row=f, column=index).value)
                dict_s[name].append(data)
            index += 1
            next_sheet_cell = sheet.cell(row=1, column=index).value
    return dict_s
